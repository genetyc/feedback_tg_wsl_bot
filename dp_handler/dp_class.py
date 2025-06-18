import asyncpg
from openpyxl import load_workbook
from aiogram.types import FSInputFile
from os import environ, remove

class Database:
    admins = [int(admin_id) for admin_id in environ["ADMINS"].split(',')]
    def __init__(self, host, port, user, password, database):
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.database = database
        self.pool = None

    async def connect(self):
        try:            
            dsn = f"postgresql://{self.user}:{self.password}@{self.host}:{self.port}/{self.database}"
            self.pool = await asyncpg.create_pool(dsn=dsn)
            print("✅ Database connection established")
        except Exception as e:
            print(f"❌ Database connection failed: {e}")

    async def close(self):
        if self.pool:
            await self.pool.close()
            print("Database pool closed")

    async def execute(self, query: str, *args):
        """Выполняет SQL-запрос без возврата данных (INSERT, UPDATE, DELETE)"""
        async with self.pool.acquire() as conn:
            await conn.execute(query, *args)

    async def fetch(self, query: str, *args):
        """Выполняет SELECT-запрос и возвращает данные"""
        async with self.pool.acquire() as conn:
            return await conn.fetch(query, *args)

    async def fetchval(self, query: str, *args):
        """Выполняет SELECT-запрос и возвращает значение"""
        async with self.pool.acquire() as conn:
            return await conn.fetchval(query, *args)

    async def update(self, telegram_id: int, column: str, answer: str, minisurvey=False):
        """Сохраняет ответ пользователя в таблице"""
        query = f"UPDATE {'public.mini_survey' if minisurvey else 'public.survey'} SET {column} = $1 WHERE telegram_id = $2"
        await self.execute(query, answer, telegram_id)

    async def add_user(self, telegram_id: int, table: str):
        """Добавляет telegram_id в таблицу survey, если его еще нет"""
        query = f'INSERT INTO {table} (telegram_id) VALUES ($1) ON CONFLICT (telegram_id) DO NOTHING'
        await self.execute(query, telegram_id)

    async def del_user(self, telegram_id: int, minisurvey=False):
        query = f'DELETE FROM public.{'survey' if not minisurvey else 'mini_survey'} WHERE telegram_id = $1'
        await self.execute(query, telegram_id)

    async def count_users(self):
        query = 'SELECT COUNT(*) FROM public.mini_survey WHERE is_complete = TRUE'
        return await self.fetchval(query)

    async def get_average_edu_rating(self):
        """Средний рейтинг и количество рецендентов из небольшого опроса"""
        rating_query = f'SELECT AVG(edu_rating) FROM public.mini_survey WHERE edu_rating > 0'
        total_query = f'SELECT COUNT(edu_rating) FROM public.mini_survey WHERE edu_rating > 0'
        
        rating_result = await self.fetchval(rating_query)
        total_result = await self.fetchval(total_query)
        return [rating_result, total_result]

    async def get_rating_range(self):
        neg_rating = await self.fetchval('SELECT COUNT(*) FROM public.mini_survey WHERE edu_rating > 0 AND edu_rating < 5')
        neutral_rating = await self.fetchval('SELECT COUNT(*) FROM public.mini_survey WHERE edu_rating > 4 AND edu_rating < 8')
        pos_rating = await self.fetchval('SELECT COUNT(*) FROM public.mini_survey WHERE edu_rating > 7')
        return [neg_rating, neutral_rating, pos_rating]

    async def gather_opinion(self):
        """Количество недовольных людей"""
        disappointed_query = f'SELECT COUNT(*) FROM public.mini_survey WHERE is_disappointed = TRUE AND is_complete = TRUE'
        return await self.fetchval(disappointed_query)
    
    async def who_would_share(self):
        """Количество людей, готовых порекомендовать студию"""
        share_query = 'SELECT COUNT(*) FROM public.mini_survey WHERE NOT would_u_share = $1 AND is_complete = True'
        return await self.fetchval(share_query, 'Нет')
    
    async def get_columns(self, table_name: str):
        async with self.pool.acquire() as conn:
            stmt = await conn.prepare(f"SELECT * FROM {table_name} LIMIT 1")
            return [attr.name for attr in stmt.get_attributes()]
    
    async def clear_table(self, exceptions=admins, case=3): # case 1: clear Survey, case 2: clear MiniSurvey, case 3: clear both
        placeholders = ','.join(f"${i+1}" for i in range(len(exceptions)))
        query_survey = f"""
            DELETE FROM public.survey
            WHERE telegram_id NOT IN ({placeholders})
        """
        query_mini_survey = f"""
            DELETE FROM public.mini_survey
            WHERE telegram_id NOT IN ({placeholders})
        """
        if case==1:
            await self.execute(query_survey, *exceptions)
        elif case==2:
            await self.execute(query_mini_survey, *exceptions)
        else:
            await self.execute(query_survey, *exceptions)
            await self.execute(query_mini_survey, *exceptions)

    from aiogram import Bot
    async def export_to_excel_and_send(self, bot: Bot, chat_id: int):
        survey_query = "SELECT * FROM public.survey"
        minis_query = "SELECT * FROM public.mini_survey"
        survey_rows = await self.fetch(survey_query)
        minis_rows = await self.fetch(minis_query)
        wb = load_workbook("templ_v2.xlsx") # `templ.xlsx` is deprecated
        ws = wb['Большой опрос']
        mini_ws = wb['Качество обучения']
        
        async def survey_export(ws, rows):
            desired_order = [
                'telegram_id', 'is_complete', 'complete_date', 'is_disappointed', 'anon','gender', 'age', 'education',
                'thanks_teacher', 'did_get_parents_report', 'report_was_good', 'report_good', 'report_bad',
                'report_is_missing', 'where_you_found_out', 'what_is_quality_edu', 'whats_your_goal',
                'education_quality', 'your_thoughts', 'how_effective', 'how_effective_2',
                'best_of_all', 'teacher_student', 'professionalism', 'difficulties', 'whats_good',
                'whats_bad', 'your_wishes'
            ]
            column_names = await self.get_columns("public.survey")
            for idx, row in enumerate(rows, start=3):
                flat_row = []
                for col in desired_order:
                    if col in column_names:
                        value = row[column_names.index(col)]
                        if isinstance(value, str) and ";" in value:
                            flat_row.extend(value.split(";"))
                        else:
                            flat_row.append(value)
                    else:
                        flat_row.append(None)
                for col_idx, value in enumerate(flat_row, start=1):
                    ws.cell(row=idx, column=col_idx, value=value)

        async def minisurvey_export(ws, rows):
            for idx, row in enumerate(rows, start=2):
                for index, x in enumerate(row, 65):
                    if index == 73 or index == 74:
                        cell = ws[f'{chr(index+1)}{idx}']
                        cell.value = x
                    elif index == 75:
                        cell = ws[f'{chr(index-2)}{idx}']
                        cell.value = x
                    else:
                        cell = ws[f'{chr(index)}{idx}']
                        cell.value = x
                    avg_rating, users_total = await self.get_average_edu_rating()
            neg_rating, neutral_rating, pos_rating = await self.get_rating_range()
            for index, value in enumerate([users_total, avg_rating, neg_rating, neutral_rating, pos_rating], start=1):
                cell = ws[f'P{index}']
                cell.value = value

        await survey_export(ws, survey_rows)
        await minisurvey_export(mini_ws, minis_rows)
        wb.save('poll.xlsx')
        file = FSInputFile('poll.xlsx')
        await bot.send_document(chat_id=chat_id, document=file, caption="📄 Данные опроса")
        remove('poll.xlsx')